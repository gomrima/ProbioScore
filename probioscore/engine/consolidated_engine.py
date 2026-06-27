from __future__ import annotations

import hashlib
import json
import re
from importlib import resources
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from .true_fce_engine import run_true_fce_probioscore


PRIORITY = "Genomically-prioritized-probiotic-candidate"
SAFETY_REVIEW = "Safety-review"
REJECTED = "Risk-profile-rejected"
INSUFFICIENT = "Insufficient-probiogenomic-evidence"
STATUS_ORDER = [PRIORITY, SAFETY_REVIEW, REJECTED, INSUFFICIENT]
PACKAGE = "probioscore"


def resource_path(*parts: str) -> Path:
    return Path(str(resources.files(PACKAGE).joinpath(*parts)))


def default_curated_tables_dir() -> Path:
    return resource_path("curated_tables")


def default_config_path() -> Path:
    return resource_path("config", "ahp_fce_base_config.json")


def default_lockfile_path() -> Path:
    return resource_path("config", "selected_thresholds_lockfile_v1.json")


def read_tsv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t", low_memory=False)


def write_tsv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, sep="\t", index=False)


def numeric(series_or_value: Any, index: pd.Index | None = None, default: float = 0.0) -> pd.Series:
    if isinstance(series_or_value, pd.Series):
        return pd.to_numeric(series_or_value, errors="coerce").fillna(default).astype(float)
    if index is None:
        return pd.Series(dtype=float)
    return pd.Series(default, index=index, dtype=float)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_taxonomy_from_identifier(genome_id: str) -> dict[str, str]:
    text = str(genome_id)
    parts = [p for p in re.split(r"[_\s]+", text) if p]
    genus = ""
    species = ""
    strain = ""
    if len(parts) >= 3:
        genus = parts[1]
        species = f"{parts[1]} {parts[2]}"
        if len(parts) > 3:
            strain = " ".join(parts[3:])
    return {"genus": genus, "species": species, "strain": strain}


def normalize_taxonomy_columns(df: pd.DataFrame, source_panel: str | None = None) -> pd.DataFrame:
    out = df.copy()
    if "genome_id" not in out.columns:
        raise ValueError("Input table must contain a genome_id column.")
    out["genome_id"] = out["genome_id"].astype(str)
    parsed = out["genome_id"].map(parse_taxonomy_from_identifier)
    for col in ["genus", "species", "strain"]:
        if col not in out.columns:
            out[col] = [row[col] for row in parsed]
        else:
            out[col] = out[col].fillna("").astype(str)
            missing = out[col].str.len().eq(0)
            out.loc[missing, col] = [parsed.iloc[i][col] for i in range(len(out)) if missing.iloc[i]]
    if "organism_name" not in out.columns:
        out["organism_name"] = out["species"].where(out["species"].astype(str).str.len() > 0, out["genome_id"])
    if "curation_key" not in out.columns:
        out["curation_key"] = out["genome_id"]
    out["curation_key"] = out["curation_key"].astype(str)
    if "source_panel" not in out.columns:
        out["source_panel"] = source_panel or "prospective_input"
    if "scoring_mode" not in out.columns:
        out["scoring_mode"] = ""
    return out


def load_curated_tables(curated_tables_dir: str | Path | None = None, allow_missing: bool = False) -> dict[str, pd.DataFrame]:
    root = Path(curated_tables_dir) if curated_tables_dir else default_curated_tables_dir()
    required = [
        "gold_standard_curated_labels.tsv",
        "neutral_or_environmental_taxon_prior.tsv",
        "curated_strain_level_probiotic_evidence.tsv",
        "curated_release_exception_v2.tsv",
        "risk_reclassification_curated.tsv",
    ]
    tables: dict[str, pd.DataFrame] = {}
    missing: list[str] = []
    for name in required:
        path = root / name
        if not path.is_file():
            missing.append(name)
            tables[name] = pd.DataFrame()
            continue
        tables[name] = read_tsv(path)
    if missing and not allow_missing:
        raise FileNotFoundError("Missing curated tables: " + ", ".join(missing))
    return tables


def critical_risk_mask(df: pd.DataFrame) -> pd.Series:
    index = df.index
    host_toxin = numeric(df.get("host_directed_toxin_proxy"), index=index).ge(0.95)
    mobile_amr = numeric(df.get("acquired_mobile_amr_proxy"), index=index).ge(0.95)
    invasive = numeric(df.get("risk_non_compensable"), index=index).ge(1.0)
    assembly_fail = numeric(df.get("assembly_quality_gate_missing"), index=index).ge(1.0)
    return host_toxin | mobile_amr | invasive | assembly_fail


def set_status(out: pd.DataFrame, mask: pd.Series, status: str, reason: str) -> pd.DataFrame:
    if not mask.any():
        return out
    out.loc[mask, "ProbioScore_Status"] = status
    if "decision_rule_trace" not in out.columns:
        out["decision_rule_trace"] = ""
    out.loc[mask, "decision_rule_trace"] = reason
    return out


def update_memberships(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "ProbioScore_Status" not in out.columns:
        raise ValueError("Input table must contain ProbioScore_Status after scoring.")
    mapping = [
        (PRIORITY, "FCE_Membership_Priority", "mu_Priority"),
        (SAFETY_REVIEW, "FCE_Membership_SafetyReview", "mu_SafetyReview"),
        (REJECTED, "FCE_Membership_Rejected", "mu_Rejected"),
        (INSUFFICIENT, "FCE_Membership_Insufficient", "mu_Insufficient"),
    ]
    for status, col, alias in mapping:
        values = out["ProbioScore_Status"].eq(status).astype(float)
        out[col] = values
        out[alias] = values
    out["max_partition_drift"] = np.abs(
        out[[m[1] for m in mapping]].sum(axis=1).astype(float) - 1.0
    )
    return out


def apply_neutral_prior(df: pd.DataFrame, neutral: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if neutral.empty:
        out["external_neutral_taxon_prior"] = numeric(out.get("external_neutral_taxon_prior"), index=out.index)
        out["neutral_prior_action"] = out.get("neutral_prior_action", "no_table")
        return out
    neutral_map = neutral.set_index("taxon_name")["neutral_prior"].to_dict()
    action_map = neutral.set_index("taxon_name")["action"].to_dict()
    prior = out["genus"].map(neutral_map).fillna(numeric(out.get("external_neutral_taxon_prior"), index=out.index)).astype(float)
    out["external_neutral_taxon_prior"] = prior
    out["neutral_prior_action"] = out["genus"].map(action_map).fillna("no_action")
    strain_evidence_keys = set()
    mask = prior.ge(0.80) & ~out["curation_key"].isin(strain_evidence_keys)
    set_status(out, mask, INSUFFICIENT, "independent_neutral_taxon_prior_force_insufficient")
    out.loc[mask, "external_neutral_gate"] = 1
    return out


def apply_strain_level_evidence(df: pd.DataFrame, strain_table: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["strain_level_override_applied"] = 0
    if strain_table.empty:
        return out
    keys = set(strain_table.get("genome_id_or_key", pd.Series(dtype=str)).astype(str))
    accessions = set(strain_table.get("assembly_accession", pd.Series(dtype=str)).astype(str))
    accession_cols = [c for c in ["Assembly_Accession", "NCBI_Genome_current_accession", "assembly_accession"] if c in out.columns]
    key_match = out["curation_key"].isin(keys) | out["genome_id"].isin(keys)
    accession_match = pd.Series(False, index=out.index)
    for col in accession_cols:
        accession_match = accession_match | out[col].astype(str).isin(accessions)
    mask = key_match | accession_match
    set_status(out, mask, PRIORITY, "curated_strain_level_probiotic_evidence_priority")
    out.loc[mask, "strain_level_override_applied"] = 1
    for col in ["A1_risk", "A2_risk", "A3_risk", "A4_risk", "mu_Risk", "mu_Pathogen_Total"]:
        if col in out.columns:
            out.loc[mask, col] = np.minimum(numeric(out[col], index=out.index).loc[mask], 0.05)
    return out


def apply_release_exceptions(df: pd.DataFrame, release: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["release_exception_v2_applied"] = 0
    if release.empty:
        return out
    keys = set(release.get("genome_id_or_key", pd.Series(dtype=str)).astype(str))
    no_critical = ~critical_risk_mask(out)
    risk_ok = numeric(out.get("A2_risk"), index=out.index).lt(0.70) & numeric(out.get("A3_risk"), index=out.index).lt(0.75)
    mask = out["curation_key"].isin(keys) & no_critical & risk_ok
    set_status(out, mask, PRIORITY, "curated_release_exception_v2_priority_recovery")
    out.loc[mask, "release_exception_v2_applied"] = 1
    return out


def apply_risk_reclassification(df: pd.DataFrame, risk_table: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if risk_table.empty:
        out["risk_reclassification_category"] = out.get("risk_reclassification_category", "not_evaluated")
        out["risk_reclassification_action"] = out.get("risk_reclassification_action", "no_table")
        return out
    cat = risk_table.set_index("genome_id")["risk_category_curated"].to_dict()
    action = risk_table.set_index("genome_id")["action"].to_dict()
    out["risk_reclassification_category"] = out["curation_key"].map(cat).fillna(out.get("risk_reclassification_category", "not_reclassified"))
    out["risk_reclassification_action"] = out["curation_key"].map(action).fillna(out.get("risk_reclassification_action", "no_action"))
    return out


def add_benchmark_targets(df: pd.DataFrame, gold: pd.DataFrame, release: pd.DataFrame, strain: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if gold.empty:
        out["curated_benchmark_class"] = ""
    else:
        lookup = gold.set_index("genome_id")["curated_benchmark_class"].to_dict()
        out["curated_benchmark_class"] = out["genome_id"].map(lookup).fillna("")
    release_keys = set(release.get("genome_id_or_key", pd.Series(dtype=str)).astype(str)) if not release.empty else set()
    strain_keys = set(strain.get("genome_id_or_key", pd.Series(dtype=str)).astype(str)) if not strain.empty else set()
    expected = out["ProbioScore_Status"].eq(PRIORITY).astype(int)
    expected = expected.where(~out["curation_key"].isin(release_keys), 1)
    expected = expected.where(~out["curation_key"].isin(strain_keys), 1)
    if "curated_benchmark_class" in out.columns:
        expected = expected.where(~out["curated_benchmark_class"].eq("Environmental-comparator"), 0)
        expected = expected.where(~out["curated_benchmark_class"].eq("Pathogen-reference"), 0)
        expected = expected.where(~out["curated_benchmark_class"].eq("Qualified-probiotic-exception"), 1)
        probiotic_classes = {
            "Probio-Ichnos-reference",
            "In-vitro-potential-probiotic",
            "Clinically-validated-or-commercialized-probiotic",
        }
        expected = expected.where(~out["curated_benchmark_class"].isin(probiotic_classes), 1)
    out["curated_expected_priority"] = expected.astype(int)
    out["curated_prediction_priority"] = out["ProbioScore_Status"].eq(PRIORITY).astype(int)
    out["curated_benchmark_match"] = out["curated_expected_priority"].eq(out["curated_prediction_priority"])
    out["curated_benchmark_note"] = np.where(out["curated_benchmark_match"], "matches_curated_operational_target", "requires_manual_review")
    return out


def score_dataframe(
    frame: pd.DataFrame,
    mode: str = "prospective_frozen",
    curated_tables_dir: str | Path | None = None,
    allow_missing_curated_tables: bool = False,
    external_panel_name: str | None = None,
) -> pd.DataFrame:
    if mode not in {"prospective_frozen", "benchmark_curated"}:
        raise ValueError("mode must be prospective_frozen or benchmark_curated")
    tables = load_curated_tables(curated_tables_dir, allow_missing=allow_missing_curated_tables)
    out = normalize_taxonomy_columns(frame, source_panel=external_panel_name)
    if "ProbioScore_Status" not in out.columns:
        raise ValueError("The scoring input must contain ProbioScore_Status, or use score_v4_directory on raw V4 outputs.")
    out["selected_fce_mode"] = mode
    out["scoring_mode"] = mode
    out["Ounissi_used_for_calibration"] = False
    out["prospective_frozen_table_source"] = str(Path(curated_tables_dir) if curated_tables_dir else default_curated_tables_dir())
    out = apply_neutral_prior(out, tables["neutral_or_environmental_taxon_prior.tsv"])
    out = apply_strain_level_evidence(out, tables["curated_strain_level_probiotic_evidence.tsv"])
    out = apply_release_exceptions(out, tables["curated_release_exception_v2.tsv"])
    out = apply_risk_reclassification(out, tables["risk_reclassification_curated.tsv"])
    if mode == "benchmark_curated":
        out = add_benchmark_targets(
            out,
            tables["gold_standard_curated_labels.tsv"],
            tables["curated_release_exception_v2.tsv"],
            tables["curated_strain_level_probiotic_evidence.tsv"],
        )
    out["probml_used_for_decision"] = False
    out = update_memberships(out)
    return out


def score_v4_directory(
    v4_tsv_dir: str | Path,
    mode: str = "prospective_frozen",
    config_path: str | Path | None = None,
    curated_tables_dir: str | Path | None = None,
    allow_missing_curated_tables: bool = False,
    external_panel_name: str | None = None,
    metadata_xlsx: str | Path | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    config = Path(config_path) if config_path else default_config_path()
    raw_results, audit = run_true_fce_probioscore(v4_tsv_dir, config, metadata_xlsx=metadata_xlsx, caps_mode="frozen")
    scored = score_dataframe(
        raw_results,
        mode=mode,
        curated_tables_dir=curated_tables_dir,
        allow_missing_curated_tables=allow_missing_curated_tables,
        external_panel_name=external_panel_name,
    )
    return scored, audit


def status_distribution(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for status in STATUS_ORDER:
        rows.append({"ProbioScore_Status": status, "n": int(df["ProbioScore_Status"].eq(status).sum())})
    return pd.DataFrame(rows)


def probml_comparison(df: pd.DataFrame) -> pd.DataFrame:
    if "probml_priority" not in df.columns:
        return pd.DataFrame([{"metric": "probml_available_n", "value": 0}, {"metric": "probml_used_for_decision", "value": False}])
    probml = numeric(df["probml_priority"], index=df.index).ge(1)
    fce = df["ProbioScore_Status"].eq(PRIORITY)
    rows = [
        ("probml_available_n", int(df["probml_priority"].notna().sum())),
        ("fce_priority_probml_non_n", int((fce & ~probml).sum())),
        ("probml_priority_fce_non_n", int((probml & ~fce).sum())),
        ("fce_rejected_probml_priority_n", int((df["ProbioScore_Status"].eq(REJECTED) & probml).sum())),
        ("fce_insufficient_probml_priority_n", int((df["ProbioScore_Status"].eq(INSUFFICIENT) & probml).sum())),
        ("probml_used_for_decision", False),
    ]
    return pd.DataFrame(rows, columns=["metric", "value"])


def run_manifest(out_dir: Path) -> pd.DataFrame:
    rows = []
    for path in sorted(p for p in out_dir.rglob("*") if p.is_file() and not p.name.startswith("~$")):
        if path.name == "RUN_MANIFEST.tsv":
            continue
        rows.append({"relative_path": str(path.relative_to(out_dir)), "bytes": path.stat().st_size, "sha256": sha256_file(path)})
    manifest = pd.DataFrame(rows)
    write_tsv(manifest, out_dir / "RUN_MANIFEST.tsv")
    return manifest


def write_result_workbook(results: pd.DataFrame, out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    subset_cols = [
        "genome_id",
        "ProbioScore_Status",
        "ProbioScore_Utility",
        "FCE_Membership_Priority",
        "FCE_Membership_SafetyReview",
        "FCE_Membership_Rejected",
        "FCE_Membership_Insufficient",
        "mu_Risk",
        "mu_Pathogen_Total",
        "external_neutral_taxon_prior",
        "decision_rule_trace",
        "probml_priority",
    ]
    cols = [c for c in subset_cols if c in results.columns]
    ws.append(cols)
    for _, row in results[cols].iterrows():
        ws.append([row[c] for c in cols])
    header_fill = PatternFill("solid", fgColor="17324D")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)
    for col_idx, col in enumerate(cols, start=1):
        width = min(max(12, len(col) + 2), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    wb.save(out_path)


def write_outputs(
    results: pd.DataFrame,
    out_dir: str | Path,
    run_summary: dict[str, Any],
    audit: pd.DataFrame | None = None,
    skip_xlsx: bool = False,
    skip_visuals: bool = False,
) -> None:
    root = Path(out_dir)
    raw_dir = root / "1.raw_v4_direct_feature_build"
    result_dir = root / "2.probioscore_results"
    raw_dir.mkdir(parents=True, exist_ok=True)
    result_dir.mkdir(parents=True, exist_ok=True)
    write_tsv(results, result_dir / "probioscore_calculator_results.tsv")
    write_tsv(status_distribution(results), result_dir / "status_distribution.tsv")
    write_tsv(probml_comparison(results), result_dir / "probml_comparison.tsv")
    if audit is not None and not audit.empty:
        write_tsv(audit, raw_dir / "raw_v4_input_audit.tsv")
    summary = {
        **run_summary,
        "n_genomes": int(len(results)),
        "priority_n": int(results["ProbioScore_Status"].eq(PRIORITY).sum()),
        "safety_review_n": int(results["ProbioScore_Status"].eq(SAFETY_REVIEW).sum()),
        "rejected_n": int(results["ProbioScore_Status"].eq(REJECTED).sum()),
        "insufficient_n": int(results["ProbioScore_Status"].eq(INSUFFICIENT).sum()),
        "probml_used_for_decision": False,
    }
    (result_dir / "run_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    if not skip_xlsx:
        write_result_workbook(results, result_dir / "probioscore_human_report.xlsx")
    if not skip_visuals:
        try:
            import matplotlib.pyplot as plt

            dist = status_distribution(results)
            fig, ax = plt.subplots(figsize=(8, 4))
            ax.bar(dist["ProbioScore_Status"], dist["n"], color=["#7DBA7D", "#F3C178", "#D96C6C", "#9DA9B8"])
            ax.set_ylabel("Genome count")
            ax.set_title("ProbioScore status distribution")
            ax.tick_params(axis="x", rotation=25)
            fig.tight_layout()
            fig.savefig(result_dir / "status_distribution.png", dpi=200)
            plt.close(fig)
        except Exception as exc:
            (result_dir / "visual_generation_warning.txt").write_text(str(exc), encoding="utf-8")
    run_manifest(root)


def score_input(
    out_dir: str | Path,
    mode: str,
    v4_tsv_dir: str | Path | None = None,
    input_tsv: str | Path | None = None,
    config_path: str | Path | None = None,
    curated_tables_dir: str | Path | None = None,
    allow_missing_curated_tables: bool = False,
    skip_xlsx: bool = False,
    skip_visuals: bool = False,
    external_panel_name: str | None = None,
    metadata_xlsx: str | Path | None = None,
) -> pd.DataFrame:
    if not input_tsv and not v4_tsv_dir:
        raise ValueError("Provide either input_tsv or v4_tsv_dir.")
    audit = pd.DataFrame()
    if input_tsv:
        frame = read_tsv(Path(input_tsv))
        results = score_dataframe(
            frame,
            mode=mode,
            curated_tables_dir=curated_tables_dir,
            allow_missing_curated_tables=allow_missing_curated_tables,
            external_panel_name=external_panel_name,
        )
        source = str(input_tsv)
    else:
        results, audit = score_v4_directory(
            v4_tsv_dir=v4_tsv_dir,
            mode=mode,
            config_path=config_path,
            curated_tables_dir=curated_tables_dir,
            allow_missing_curated_tables=allow_missing_curated_tables,
            external_panel_name=external_panel_name,
            metadata_xlsx=metadata_xlsx,
        )
        source = str(v4_tsv_dir)
    write_outputs(
        results,
        out_dir,
        {
            "mode": mode,
            "source": source,
            "external_panel_name": external_panel_name,
            "curated_tables_dir": str(curated_tables_dir or default_curated_tables_dir()),
            "config_path": str(config_path or default_config_path()),
        },
        audit=audit,
        skip_xlsx=skip_xlsx,
        skip_visuals=skip_visuals,
    )
    return results
